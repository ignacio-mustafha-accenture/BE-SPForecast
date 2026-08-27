"""
import_offerings_clients.py
---------------------------
Sincroniza offering, cliente y first_available de los 99 empleados
combinando las dos fuentes de verdad.

FUENTES Y PRECEDENCIA
---------------------
CSV  "01 - FORECAST CONTROL S&P"   -> 99 empleados. Base de todo.
Excel "Forecast Update"            -> 70 empleados de S&P. Refina al CSV.

  offering         CSV Sub-Offerings -> codigo corto.
                   El Excel pisa con "Tools" a los que marca asi,
                   porque distingue Tools dentro de Procurement Reinvention
                   y el CSV no.

  client           Excel gana (mas actualizado). El CSV cubre a los que
                   el Excel no tiene.

  first_available  Excel gana. El CSV cubre el resto.

PARSEO DE "Comentarios" EN EL CSV
---------------------------------
Convive en dos formatos:
    A)  100|Blackstone|13/07/2026|31/12/2026
        pct | cliente | roll_on | roll_off
    B)  Client:Grupo Mariposa|00/01/1900 to 30/08/2026|Next PTO 13 al 16/7
        Client:<nombre> | <rango de fechas> | <nota de PTO>

El cargador original tomo el tramo de fechas como si fuera el cliente.
De ahi que la columna client tenga valores como '16/07/2026 to 15/10/2026'.

USO
---
    python import_offerings_clients.py <csv> <xlsx>            # dry-run
    python import_offerings_clients.py <csv> <xlsx> --apply    # escribe
"""

import argparse
import asyncio
import csv
import datetime
import sys
from collections import Counter

sys.path.insert(0, ".")

import asyncpg
import openpyxl

from app.config import settings

# --- CSV ---
CSV_HEADER_ROW = 4          # 0-based: fila 5 del archivo
CSV_FIRST_DATA_ROW = 5
CSV_COL = {
    "status":        0,
    "offering":      1,
    "sub_offering":  2,
    "comentarios":   8,
    "first_avail":   9,
    "level":        10,
    "eid":          11,
}

# --- Excel ---
XL_SHEET = "Forecast Update"
XL_FIRST_DATA_ROW = 8
XL_COL = {
    "eid": 1, "offering_label": 2, "cliente": 4, "status": 5, "first_available": 12,
}
XL_JUNK = ("assumptions", "cascadeo", "lista de", "feriados")

# Sub-Offering del CSV -> codigo corto de la taxonomia
SUB_OFFERING_MAP = {
    "Spend Optimization":      "SO",
    "Procurement Reinvention": "PR",
    "S4":                      "S4",
    "Ariba":                   "Ariba",
    "Oracle":                  "Oracle",
}

# Offering Label del Excel que refina al CSV
XL_OFFERING_OVERRIDE = {"Tools": "Tools"}

VALID_OFFERINGS = {"SO", "PR", "Tools", "S4", "Ariba", "Oracle"}


def clean(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    return s or None


def parse_ddmmyyyy(s):
    """'13/07/2026' -> date. Devuelve None si es invalida o centinela."""
    s = (s or "").strip()
    if not s or s.startswith("00/"):
        return None
    try:
        return datetime.datetime.strptime(s, "%d/%m/%Y").date()
    except ValueError:
        return None


def parse_comentarios(raw):
    """
    Extrae el nombre del cliente de la columna Comentarios del CSV.
    Devuelve None si no puede determinarlo.
    """
    raw = (raw or "").strip()
    if not raw:
        return None
    partes = [p.strip() for p in raw.split("|")]

    # Formato A: pct | cliente | roll_on | roll_off
    if len(partes) >= 4 and partes[0].replace(".", "").isdigit():
        return partes[1] or None

    # Formato B: Client:<nombre> | ... | ...
    if partes[0].lower().startswith("client:"):
        return partes[0].split(":", 1)[1].strip() or None

    return None


def read_csv(path):
    with open(path, encoding="utf-8-sig", errors="replace") as f:
        rows = list(csv.reader(f))

    out = {}
    for r in rows[CSV_FIRST_DATA_ROW:]:
        if len(r) <= CSV_COL["eid"]:
            continue
        eid = r[CSV_COL["eid"]].strip()
        if not eid:
            continue
        sub = r[CSV_COL["sub_offering"]].strip()
        out[eid] = {
            "sub_offering":    sub,
            "offering":        SUB_OFFERING_MAP.get(sub),
            "client":          parse_comentarios(r[CSV_COL["comentarios"]]),
            "first_available": parse_ddmmyyyy(r[CSV_COL["first_avail"]]),
            "status":          r[CSV_COL["status"]].strip() or None,
        }
    return out


def read_excel(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if XL_SHEET not in wb.sheetnames:
        raise SystemExit(f"El Excel no tiene la hoja '{XL_SHEET}'")
    ws = wb[XL_SHEET]

    out = {}
    for raw in ws.iter_rows(min_row=XL_FIRST_DATA_ROW, values_only=True):
        eid = clean(raw[XL_COL["eid"]]) if len(raw) > XL_COL["eid"] else None
        if not eid or any(eid.lower().startswith(j) for j in XL_JUNK):
            continue

        def get(k):
            i = XL_COL[k]
            return clean(raw[i]) if len(raw) > i else None

        if not (get("offering_label") or get("cliente")):
            continue

        out[eid] = {
            "offering_label":  get("offering_label"),
            "client":          get("cliente"),
            "first_available": get("first_available"),
            "status":          get("status"),
        }
    wb.close()
    return out


def merge(csv_data, xl_data, alias):
    """
    Combina ambas fuentes aplicando la precedencia documentada arriba.
    'alias' mapea EID-del-Excel -> EID-real, para los casos con typo.
    """
    merged = {}

    for eid, c in csv_data.items():
        merged[eid] = {
            "offering":        c["offering"],
            "client":          c["client"],
            "first_available": c["first_available"],
            "status":          c["status"],
            "fuente_offering": "csv" if c["offering"] else None,
            "fuente_client":   "csv" if c["client"] else None,
        }

    for raw_eid, x in xl_data.items():
        eid = alias.get(raw_eid, raw_eid)
        m = merged.setdefault(eid, {
            "offering": None, "client": None, "first_available": None,
            "status": None, "fuente_offering": None, "fuente_client": None,
        })

        # El Excel solo pisa el offering cuando aporta granularidad que el CSV no tiene
        override = XL_OFFERING_OVERRIDE.get(x["offering_label"] or "")
        if override:
            m["offering"] = override
            m["fuente_offering"] = "excel"

        if x["client"]:
            m["client"] = x["client"]
            m["fuente_client"] = "excel"
        if x["first_available"]:
            m["first_available"] = x["first_available"]
        if x["status"]:
            m["status"] = x["status"]

    return merged


async def build_alias(conn, xl_data):
    """Resuelve EIDs del Excel que vienen como nombre o con typo."""
    db = await conn.fetch("SELECT eid, name FROM employees")
    by_eid = {r["eid"] for r in db}
    by_name = {(r["name"] or "").strip().lower(): r["eid"] for r in db}

    alias, sin_match = {}, []
    for e in xl_data:
        if e in by_eid:
            continue
        hit = by_name.get(e.lower())
        if hit:
            alias[e] = hit
        else:
            sin_match.append(e)
    return alias, sin_match


async def main(csv_path, xlsx_path, apply):
    csv_data = read_csv(csv_path)
    xl_data = read_excel(xlsx_path)
    print(f"CSV:    {len(csv_data)} empleados")
    print(f"Excel:  {len(xl_data)} empleados\n")

    conn = await asyncpg.connect(
        host=settings.DB_HOST, port=settings.DB_PORT,
        user=settings.DB_USER, password=settings.DB_PASSWORD,
        database=settings.DB_NAME, ssl="require",
    )

    alias, sin_match = await build_alias(conn, xl_data)
    if alias:
        print("EIDs del Excel resueltos por nombre:")
        for k, v in alias.items():
            print(f"   {k!r} -> {v}")
        print()
    if sin_match:
        print(f"EIDs del Excel que no existen en la DB ({len(sin_match)}):")
        for m in sin_match:
            print(f"   - {m}")
        print()

    merged = merge(csv_data, xl_data, alias)

    db_eids = {r["eid"] for r in await conn.fetch("SELECT eid FROM employees")}
    aplicables = {e: d for e, d in merged.items() if e in db_eids}
    print(f"Empleados a actualizar: {len(aplicables)} de {len(db_eids)} en la DB\n")

    dist = Counter(d["offering"] for d in aplicables.values())
    print("Distribucion de offering resultante:")
    for k, v in dist.most_common():
        print(f"   {k or '(sin offering)':16} {v}")
    invalidos = {k for k in dist if k and k not in VALID_OFFERINGS}
    if invalidos:
        print(f"   AVISO: valores fuera de la taxonomia: {invalidos}")
    print()

    fuentes = Counter(d["fuente_client"] for d in aplicables.values())
    print("Origen del cliente:")
    for k, v in fuentes.most_common():
        print(f"   {k or '(sin dato)':16} {v}")
    print()

    actuales = {
        r["eid"]: r
        for r in await conn.fetch("""
            WITH latest AS (
                SELECT DISTINCT ON (eid) * FROM forecast_update
                ORDER BY eid, updated_at DESC NULLS LAST
            )
            SELECT e.eid, l.client, l.offering, l.first_available
            FROM employees e LEFT JOIN latest l ON l.eid = e.eid
        """)
    }

    cambios = Counter()
    muestras = []
    for eid, d in aplicables.items():
        cur = actuales.get(eid, {})
        diff = []
        if d["offering"] and d["offering"] != cur.get("offering"):
            cambios["offering"] += 1
            diff.append(f"offering: {cur.get('offering')!r} -> {d['offering']!r}")
        if d["client"] and d["client"] != cur.get("client"):
            cambios["client"] += 1
            diff.append(f"client: {cur.get('client')!r} -> {d['client']!r}")
        if d["first_available"] and d["first_available"] != cur.get("first_available"):
            cambios["first_available"] += 1
            diff.append(f"first_available: {cur.get('first_available')} -> {d['first_available']}")
        if diff and len(muestras) < 10:
            muestras.append((eid, diff))

    print("Cambios que se aplicarian:")
    for k, v in cambios.most_common():
        print(f"   {k:18} {v}")
    print()
    if muestras:
        print("Ejemplos:")
        for eid, diff in muestras:
            print(f"   {eid}")
            for d_ in diff:
                print(f"      {d_}")
        print()

    if not apply:
        print("[DRY RUN] No se escribio nada. Volve a correr con --apply.")
        await conn.close()
        return

    async with conn.transaction():
        await conn.execute(
            "ALTER TABLE forecast_update ADD COLUMN IF NOT EXISTS status VARCHAR(40)"
        )

        n = 0
        for eid, d in aplicables.items():
            sets, params = [], []

            def add(col, val):
                if val is None:
                    return
                params.append(val)
                sets.append(f"{col} = ${len(params)}")

            add("offering",        d["offering"])
            add("client",          d["client"])
            add("first_available", d["first_available"])
            add("status",          d["status"])

            if not sets:
                continue

            params.append(eid)
            res = await conn.execute(
                f"UPDATE forecast_update SET {', '.join(sets)}, updated_at = NOW() "
                f"WHERE eid = ${len(params)}",
                *params,
            )
            if res.endswith(" 0"):
                await conn.execute(
                    """
                    INSERT INTO forecast_update
                        (eid, offering, client, first_available, status,
                         updated_at, scenario_type)
                    VALUES ($1,$2,$3,$4,$5,NOW(),'assumption')
                    """,
                    eid, d["offering"], d["client"], d["first_available"], d["status"],
                )

            if d["offering"]:
                await conn.execute(
                    "UPDATE employees SET offering = $1 WHERE eid = $2",
                    d["offering"], eid,
                )
            n += 1

    print(f"OK. {n} empleados actualizados.")
    await conn.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("csv", help="Ruta al CSV 01 - FORECAST CONTROL S&P")
    ap.add_argument("xlsx", help="Ruta al Excel maestro de forecast")
    ap.add_argument("--apply", action="store_true", help="Escribe. Sin el flag es dry-run.")
    a = ap.parse_args()
    asyncio.run(main(a.csv, a.xlsx, a.apply))