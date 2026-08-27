"""
import_forecast_excel.py
------------------------
Sincroniza forecast_update y employees desde la hoja "Forecast Update"
del Excel maestro de S&P.

Campos que importa:
    forecast_update : offering, client, first_available, roll_on, roll_off,
                      te_approver, office, next_client, status
    employees       : cl, location, country, hire_date, offering

Reglas:
  - Match por EID exacto. Si el EID del Excel tiene formato de nombre
    ("Jose Alvarado"), intenta matchear contra employees.name.
  - NUNCA pisa un valor de la DB con un vacio del Excel. Solo actualiza
    los campos que el Excel trae con dato.
  - La columna "Days 2 Availab." NO se importa: el -99 es un centinela de
    "En reserva", no un conteo de dias. Se importa el Status y el backend
    calcula los dias desde first_available.

Uso:
    python import_forecast_excel.py <ruta_al_excel>            # dry-run
    python import_forecast_excel.py <ruta_al_excel> --apply    # escribe
"""

import argparse
import asyncio
import datetime
import sys

sys.path.insert(0, ".")

import asyncpg
import openpyxl

from app.config import settings

SHEET = "Forecast Update"
HEADER_ROW = 7
FIRST_DATA_ROW = 8

# Filas de notas al pie de la hoja que no son empleados
JUNK_PREFIXES = ("assumptions", "cascadeo", "lista de", "feriados")

# Indices de columna (0-based) segun la fila de encabezado
COL = {
    "location":        0,
    "eid":             1,
    "offering_label":  2,
    "cl":              3,
    "cliente":         4,
    "status":          5,
    "hire_date":       7,
    "office":          8,
    "te_approver":     9,
    "roll_on":        10,
    "roll_off":       11,
    "first_available": 12,
    "next_client":    15,
}

# Location del Excel -> country en la DB
LOCATION_TO_COUNTRY = {
    "ARG": "AR",
    "MX":  "MX",
    "CR":  "CR",
}


def clean(v):
    """Normaliza una celda: None si esta vacia, str/date si tiene contenido."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    s = str(v).strip()
    return s or None


def to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def read_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"El archivo no tiene la hoja '{SHEET}'")
    ws = wb[SHEET]

    rows = []
    for raw in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        eid = clean(raw[COL["eid"]] if len(raw) > COL["eid"] else None)
        if not eid:
            continue
        if any(eid.lower().startswith(p) for p in JUNK_PREFIXES):
            continue

        def get(key):
            i = COL[key]
            return clean(raw[i]) if len(raw) > i else None

        # Una fila de empleado real tiene al menos offering o cliente
        if not (get("offering_label") or get("cliente")):
            continue

        rows.append({
            "eid":             eid,
            "location":        get("location"),
            "offering_label":  get("offering_label"),
            "cl":              to_int(get("cl")),
            "cliente":         get("cliente"),
            "status":          get("status"),
            "hire_date":       get("hire_date"),
            "office":          get("office"),
            "te_approver":     get("te_approver"),
            "roll_on":         get("roll_on"),
            "roll_off":        get("roll_off"),
            "first_available": get("first_available"),
            "next_client":     get("next_client"),
        })

    wb.close()
    return rows


async def ensure_status_column(conn):
    """Agrega forecast_update.status si no existe. Idempotente."""
    await conn.execute("""
        ALTER TABLE forecast_update
            ADD COLUMN IF NOT EXISTS status VARCHAR(40)
    """)


async def resolve_eids(conn, rows: list[dict]) -> tuple[dict, list[str]]:
    """
    Devuelve (mapa excel_eid -> eid_real_en_db, lista_de_no_encontrados).
    Si el EID del Excel es en realidad un nombre, lo busca por employees.name.
    """
    db_rows = await conn.fetch("SELECT eid, name FROM employees")
    by_eid = {r["eid"] for r in db_rows}
    by_name = {(r["name"] or "").strip().lower(): r["eid"] for r in db_rows}

    resolved, missing = {}, []
    for r in rows:
        e = r["eid"]
        if e in by_eid:
            resolved[e] = e
            continue
        hit = by_name.get(e.lower())
        if hit:
            resolved[e] = hit
            continue
        missing.append(e)
    return resolved, missing


async def main(path: str, apply: bool):
    rows = read_excel(path)
    print(f"Excel: {len(rows)} filas de empleado leidas de '{SHEET}'\n")

    conn = await asyncpg.connect(
        host=settings.DB_HOST,
        port=settings.DB_PORT,
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        database=settings.DB_NAME,
        ssl="require",
    )

    if apply:
        await ensure_status_column(conn)
    else:
        # En dry-run la columna puede no existir todavia
        exists = await conn.fetchval("""
            SELECT 1 FROM information_schema.columns
            WHERE table_name='forecast_update' AND column_name='status'
        """)
        if not exists:
            print("NOTA: forecast_update.status no existe todavia. "
                  "Se creara al correr con --apply.\n")

    resolved, missing = await resolve_eids(conn, rows)
    print(f"EIDs resueltos:    {len(resolved)}")
    print(f"EIDs sin match:    {len(missing)}")
    for m in missing:
        print(f"   - {m}")
    print()

    # Estado actual en DB para comparar
    current = {
        r["eid"]: r
        for r in await conn.fetch("""
            WITH latest AS (
                SELECT DISTINCT ON (eid) *
                FROM forecast_update
                ORDER BY eid, updated_at DESC NULLS LAST
            )
            SELECT e.eid, e.cl, e.country, e.offering,
                   l.client, l.offering AS fu_offering, l.first_available,
                   l.roll_on, l.roll_off, l.te_approver, l.office
            FROM employees e
            LEFT JOIN latest l ON l.eid = e.eid
        """)
    }

    changes = {"offering": 0, "client": 0, "first_available": 0,
               "roll_on": 0, "roll_off": 0, "te_approver": 0,
               "office": 0, "cl": 0, "status": 0}
    samples = []

    for r in rows:
        eid = resolved.get(r["eid"])
        if not eid:
            continue
        cur = current.get(eid, {})
        diff = []

        if r["offering_label"] and r["offering_label"] != cur.get("fu_offering"):
            changes["offering"] += 1
            diff.append(f"offering: {cur.get('fu_offering')!r} -> {r['offering_label']!r}")
        if r["cliente"] and r["cliente"] != cur.get("client"):
            changes["client"] += 1
            diff.append(f"client: {cur.get('client')!r} -> {r['cliente']!r}")
        if r["first_available"] and r["first_available"] != cur.get("first_available"):
            changes["first_available"] += 1
            diff.append(f"first_available: {cur.get('first_available')} -> {r['first_available']}")
        if r["roll_on"] and r["roll_on"] != cur.get("roll_on"):
            changes["roll_on"] += 1
        if r["roll_off"] and r["roll_off"] != cur.get("roll_off"):
            changes["roll_off"] += 1
        if r["te_approver"] and r["te_approver"] != cur.get("te_approver"):
            changes["te_approver"] += 1
        if r["office"] and r["office"] != cur.get("office"):
            changes["office"] += 1
        if r["cl"] is not None and r["cl"] != to_int(cur.get("cl")):
            changes["cl"] += 1
        if r["status"]:
            changes["status"] += 1

        if diff and len(samples) < 12:
            samples.append((eid, diff))

    print("Cambios que se aplicarian:")
    for k, v in changes.items():
        print(f"   {k:18} {v}")
    print()

    if samples:
        print("Ejemplos:")
        for eid, diff in samples:
            print(f"   {eid}")
            for d in diff:
                print(f"      {d}")
        print()

    if not apply:
        print("[DRY RUN] No se escribio nada. Volve a correr con --apply.")
        await conn.close()
        return

    applied = 0
    async with conn.transaction():
        for r in rows:
            eid = resolved.get(r["eid"])
            if not eid:
                continue

            # --- forecast_update: solo campos con dato en el Excel ---
            sets, params = [], []

            def add(col, val):
                if val is None:
                    return
                params.append(val)
                sets.append(f"{col} = ${len(params)}")

            add("offering",        r["offering_label"])
            add("client",          r["cliente"])
            add("first_available", r["first_available"])
            add("roll_on",         r["roll_on"])
            add("roll_off",        r["roll_off"])
            add("te_approver",     r["te_approver"])
            add("office",          r["office"])
            add("next_client",     r["next_client"])
            add("status",          r["status"])

            if sets:
                params.append(eid)
                updated = await conn.execute(
                    f"""
                    UPDATE forecast_update
                    SET {', '.join(sets)}, updated_at = NOW()
                    WHERE eid = ${len(params)}
                    """,
                    *params,
                )
                # Si el empleado no tenia fila en forecast_update, la creamos
                if updated.endswith(" 0"):
                    await conn.execute(
                        """
                        INSERT INTO forecast_update
                            (eid, offering, client, first_available, roll_on,
                             roll_off, te_approver, office, next_client,
                             status, updated_at, scenario_type)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,NOW(),'assumption')
                        """,
                        eid, r["offering_label"], r["cliente"], r["first_available"],
                        r["roll_on"], r["roll_off"], r["te_approver"], r["office"],
                        r["next_client"], r["status"],
                    )

            # --- employees: atributos de la persona ---
            emp_sets, emp_params = [], []

            def add_emp(col, val):
                if val is None:
                    return
                emp_params.append(val)
                emp_sets.append(f"{col} = ${len(emp_params)}")

            add_emp("cl", r["cl"])
            add_emp("offering", r["offering_label"])
            add_emp("hire_date", r["hire_date"])
            if r["location"]:
                add_emp("country", LOCATION_TO_COUNTRY.get(r["location"], r["location"]))

            if emp_sets:
                emp_params.append(eid)
                await conn.execute(
                    f"UPDATE employees SET {', '.join(emp_sets)} WHERE eid = ${len(emp_params)}",
                    *emp_params,
                )

            applied += 1

    print(f"OK. {applied} empleados actualizados.")
    await conn.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("excel", help="Ruta al Excel maestro de forecast")
    ap.add_argument("--apply", action="store_true",
                    help="Escribe los cambios. Sin este flag es dry-run.")
    args = ap.parse_args()
    asyncio.run(main(args.excel, args.apply))