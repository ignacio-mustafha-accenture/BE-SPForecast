import argparse, asyncio, csv, datetime, sys
from collections import Counter
from pathlib import Path
sys.path.insert(0, ".")
import asyncpg
from app.config import settings

COLUMNAS = {
    "eid": ["EID"],
    "location": ["Location"],
    "offering": ["Offering Label"],
    "cl": ["CL"],
    "cliente": ["Cliente"],
    "status": ["Status"],
    "hire_date": ["Hire Date"],
    "office": ["Office"],
    "te_approver": ["T&E approver (level 7)", "T&E approver"],
    "roll_on": ["Roll-on"],
    "roll_off": ["Roll-off"],
    "first_available": ["First"],
    "next_client": ["Next Client"],
}
OBLIGATORIAS = ["eid", "offering", "cliente"]
VALID_OFFERINGS = {"SO", "PR", "Tools", "S4", "Ariba", "Oracle"}
JUNK = ("assumptions", "cascadeo", "lista de", "feriados", "total", "hc ")
LOCATION_TO_COUNTRY = {"ARG": "AR", "MX": "MX", "CR": "CR"}
FORMATOS_FECHA = ("%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y")


def norm(s):
    return " ".join(str(s or "").split()).strip().lower()


def leer_filas(path):
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        hoja = next((h for h in wb.sheetnames if norm(h) == "forecast update"), None)
        if hoja is None:
            raise SystemExit("No encuentro la hoja 'Forecast Update'. Hay: %s" % wb.sheetnames)
        filas = [["" if c is None else c for c in f] for f in wb[hoja].iter_rows(values_only=True)]
        wb.close()
        return filas
    with open(path, encoding="utf-8-sig", errors="replace") as f:
        return list(csv.reader(f))


def ubicar_encabezado(filas):
    for i, fila in enumerate(filas[:30]):
        celdas = [norm(c) for c in fila]
        if "eid" not in celdas:
            continue
        mapa = {}
        for interno, cands in COLUMNAS.items():
            for cand in cands:
                if norm(cand) in celdas:
                    mapa[interno] = celdas.index(norm(cand))
                    break
        if [c for c in OBLIGATORIAS if c not in mapa]:
            continue
        return i, mapa
    raise SystemExit("No encontre la fila de encabezado con EID, Offering Label y Cliente.")


def celda(fila, mapa, clave):
    i = mapa.get(clave)
    if i is None or i >= len(fila):
        return None
    v = fila[i]
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    s = str(v).strip()
    return s or None


def a_fecha(v):
    if v is None:
        return None
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    if not s or s.startswith("00/"):
        return None
    for fmt in FORMATOS_FECHA:
        try:
            d = datetime.datetime.strptime(s, fmt).date()
            return d if d.year >= 2000 else None
        except ValueError:
            continue
    return None


def a_entero(v):
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return None


def parsear(path):
    filas = leer_filas(path)
    h, mapa = ubicar_encabezado(filas)
    print("Encabezado en la fila %d" % (h + 1))
    print("Columnas mapeadas:")
    for k, v in sorted(mapa.items(), key=lambda x: x[1]):
        print("   %-16s -> col %d" % (k, v))
    falt = [k for k in COLUMNAS if k not in mapa]
    if falt:
        print("No estan en el archivo (se ignoran): %s" % ", ".join(falt))
    print()
    out = []
    for fila in filas[h + 1:]:
        eid = celda(fila, mapa, "eid")
        if not eid:
            continue
        eid = str(eid).strip()
        if any(eid.lower().startswith(j) for j in JUNK):
            continue
        if not (celda(fila, mapa, "offering") or celda(fila, mapa, "cliente")):
            continue
        out.append({
            "eid": eid,
            "location": celda(fila, mapa, "location"),
            "offering": celda(fila, mapa, "offering"),
            "cl": a_entero(celda(fila, mapa, "cl")),
            "cliente": celda(fila, mapa, "cliente"),
            "status": celda(fila, mapa, "status"),
            "hire_date": a_fecha(celda(fila, mapa, "hire_date")),
            "office": celda(fila, mapa, "office"),
            "te_approver": celda(fila, mapa, "te_approver"),
            "roll_on": a_fecha(celda(fila, mapa, "roll_on")),
            "roll_off": a_fecha(celda(fila, mapa, "roll_off")),
            "first_available": a_fecha(celda(fila, mapa, "first_available")),
            "next_client": celda(fila, mapa, "next_client"),
        })
    return out


async def resolver_eids(conn, filas):
    db = await conn.fetch("SELECT eid, name FROM employees")
    por_eid = {r["eid"] for r in db}
    por_nombre = {(r["name"] or "").strip().lower(): r["eid"] for r in db}
    alias, sin_match = {}, []
    for f in filas:
        e = f["eid"]
        if e in por_eid:
            alias[e] = e
        elif e.lower() in por_nombre:
            alias[e] = por_nombre[e.lower()]
        else:
            sin_match.append(e)
    return alias, sin_match


async def main(path, apply):
    filas = parsear(path)
    print("%d filas de empleado leidas" % len(filas))
    print()
    print("Offering en el archivo:")
    for k, v in Counter(f["offering"] for f in filas).most_common():
        marca = "" if k in VALID_OFFERINGS else "   <-- fuera de la taxonomia"
        print("   %-10s %d%s" % (k, v, marca))
    print()
    conn = await asyncpg.connect(host=settings.DB_HOST, port=settings.DB_PORT,
        user=settings.DB_USER, password=settings.DB_PASSWORD,
        database=settings.DB_NAME, ssl="require")
    alias, sin_match = await resolver_eids(conn, filas)
    print("EIDs que matchean: %d" % len(alias))
    if sin_match:
        print("EIDs sin match en la base (%d):" % len(sin_match))
        for m in sin_match:
            print("   - %s" % m)
    print()
    actuales = {r["eid"]: r for r in await conn.fetch("""
        WITH latest AS (SELECT DISTINCT ON (eid) * FROM forecast_update
                        ORDER BY eid, updated_at DESC NULLS LAST)
        SELECT e.eid, l.client, l.offering, l.first_available,
               l.roll_on, l.roll_off, l.status
        FROM employees e LEFT JOIN latest l ON l.eid = e.eid""")}
    cambios, muestras = Counter(), []
    for f in filas:
        eid = alias.get(f["eid"])
        if not eid:
            continue
        cur = actuales.get(eid, {})
        diff = []
        for campo, col in [("offering","offering"),("cliente","client"),
                           ("first_available","first_available"),("roll_on","roll_on"),
                           ("roll_off","roll_off"),("status","status")]:
            nuevo = f[campo]
            if nuevo is not None and nuevo != cur.get(col):
                cambios[campo] += 1
                if campo in ("offering","cliente","roll_off"):
                    diff.append("%s: %r -> %r" % (campo, cur.get(col), nuevo))
        if diff and len(muestras) < 10:
            muestras.append((eid, diff))
    print("Cambios que se aplicarian:")
    for k, v in cambios.most_common():
        print("   %-18s %d" % (k, v))
    print()
    if muestras:
        print("Ejemplos:")
        for eid, diff in muestras:
            print("   %s" % eid)
            for d in diff:
                print("      %s" % d)
        print()
    if not apply:
        print("[DRY RUN] No se escribio nada. Volve a correr con --apply.")
        await conn.close()
        return
    n = 0
    async with conn.transaction():
        await conn.execute("ALTER TABLE forecast_update ADD COLUMN IF NOT EXISTS status VARCHAR(40)")
        for f in filas:
            eid = alias.get(f["eid"])
            if not eid:
                continue
            sets, params = [], []
            def add(col, val):
                if val is None:
                    return
                params.append(val)
                sets.append("%s = $%d" % (col, len(params)))
            add("offering", f["offering"]); add("client", f["cliente"])
            add("first_available", f["first_available"]); add("roll_on", f["roll_on"])
            add("roll_off", f["roll_off"]); add("te_approver", f["te_approver"])
            add("office", f["office"]); add("next_client", f["next_client"])
            add("status", f["status"])
            if sets:
                params.append(eid)
                res = await conn.execute("UPDATE forecast_update SET %s, updated_at = NOW() WHERE eid = $%d"
                                         % (", ".join(sets), len(params)), *params)
                if res.endswith(" 0"):
                    await conn.execute("""INSERT INTO forecast_update
                        (eid, offering, client, first_available, roll_on, roll_off,
                         te_approver, office, next_client, status, updated_at, scenario_type)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,NOW(),'assumption')""",
                        eid, f["offering"], f["cliente"], f["first_available"], f["roll_on"],
                        f["roll_off"], f["te_approver"], f["office"], f["next_client"], f["status"])
            es, ep = [], []
            def add_emp(col, val):
                if val is None:
                    return
                ep.append(val)
                es.append("%s = $%d" % (col, len(ep)))
            add_emp("offering", f["offering"]); add_emp("cl", f["cl"])
            add_emp("hire_date", f["hire_date"])
            if f["location"]:
                add_emp("country", LOCATION_TO_COUNTRY.get(f["location"], f["location"]))
            if es:
                ep.append(eid)
                await conn.execute("UPDATE employees SET %s WHERE eid = $%d" % (", ".join(es), len(ep)), *ep)
            n += 1
    print("OK. %d empleados actualizados." % n)
    await conn.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("archivo")
    ap.add_argument("--apply", action="store_true")
    a = ap.parse_args()
    asyncio.run(main(Path(a.archivo), a.apply))
