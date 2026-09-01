"""
load_from_excel.py — Lector directo del Excel con colores

Lee el .xlsx con openpyxl (data_only=True) para preservar la info de color
de celda, que es el mecanismo que distingue tipos de horas:

  Sin color / 00000000  → Hard Lock (chg_hl)
  FFFFEB9C  amarillo    → Assumptions ISG Assessment (chg_sl)
  FFFFCCCC  rosado      → Assumptions No R (chg_sl)
  FFC0E6F5 / theme:4    → Assumptions R (chg_sl)
  FFFFC7CE  salmón      → Variante No R (chg_sl)
  FFCC66FF  violeta     → Cascadeo de hs (chg_cascadeadas)

Uso:
  python scripts/load_from_excel.py [path/to/forecast.xlsx] [--dry-run]

Si no se pasa path, busca el primer *.xlsx en el directorio actual.
"""

import asyncio
import logging
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, date
from pathlib import Path

sys.path.insert(0, '.')
from app.config import settings
import asyncpg

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s: %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constantes de layout del Excel
# ---------------------------------------------------------------------------
SHEET_NAME = None   # None = primera hoja
DATA_ROW_START = 11
DATA_ROW_END = 91

# Columnas de empleado (1-indexed, estilo openpyxl)
COL_LOCATION = 1     # ARG / MX / CR
COL_EID = 2          # Enterprise ID
COL_OFFERING_LABEL = 3
COL_OFFERING_REPORT = 4
COL_CL = 5           # nivel numérico (ej: 7, 8, 9)
COL_CLIENT = 6       # cliente actual
COL_STATUS = 7       # status (NJ / Active / etc.)
COL_ROLL_ON = 15
COL_ROLL_OFF = 16
COL_FIRST_AVAIL = 17
COL_NEXT_CLIENT = 20

# Mapeo location Excel → (country, location)
LOCATION_MAP = {
    'ARG': ('AR', 'AR'),
    'MX':  ('MX', 'MX'),
    'CR':  ('CR', 'CR'),
}

# ---------------------------------------------------------------------------
# Colores de celda CHG
# ---------------------------------------------------------------------------
CASCADE_COLORS = {'FFCC66FF'}

SL_COLORS = {
    'FFFFEB9C',   # amarillo — Assumptions ISG Assessment
    'FFFFCCCC',   # rosado   — Assumptions No R
    'FFC0E6F5',   # celeste  — Assumptions R
    'FFFFC7CE',   # salmón   — Variante No R
    'FF44B3E1',   # celeste oscuro (variante)
    'FF00B0F0',   # azul claro
}

# ---------------------------------------------------------------------------
# Columnas de períodos (1-indexed)
# Cada período: chg (CHG total), sah (SAH), chg_pct (CHG%)
# ---------------------------------------------------------------------------
PERIOD_COLUMNS = {
    # FY26
    'Sep-P1': {'chg': 21, 'sah': 22, 'chg_pct': 23},
    'Sep-P2': {'chg': 24, 'sah': 25, 'chg_pct': 26},
    'Oct-P1': {'chg': 27, 'sah': 28, 'chg_pct': 29},
    'Oct-P2': {'chg': 30, 'sah': 31, 'chg_pct': 32},
    'Nov-P1': {'chg': 33, 'sah': 34, 'chg_pct': 35},
    'Nov-P2': {'chg': 36, 'sah': 37, 'chg_pct': 38},
    'Dic-P1': {'chg': 39, 'sah': 40, 'chg_pct': 41},
    'Dic-P2': {'chg': 42, 'sah': 43, 'chg_pct': 44},
    'Ene-P1': {'chg': 45, 'sah': 46, 'chg_pct': 47},
    'Ene-P2': {'chg': 48, 'sah': 49, 'chg_pct': 50},
    'Feb-P1': {'chg': 51, 'sah': 52, 'chg_pct': 53},
    'Feb-P2': {'chg': 54, 'sah': 55, 'chg_pct': 56},
    'Mar-P1': {'chg': 57, 'sah': 58, 'chg_pct': 59},
    'Mar-P2': {'chg': 60, 'sah': 61, 'chg_pct': 62},
    'Abr-P1': {'chg': 63, 'sah': 64, 'chg_pct': 65},
    'Abr-P2': {'chg': 66, 'sah': 67, 'chg_pct': 68},
    'May-P1': {'chg': 69, 'sah': 70, 'chg_pct': 71},
    'May-P2': {'chg': 72, 'sah': 73, 'chg_pct': 74},
    'Jun-P1': {'chg': 75, 'sah': 76, 'chg_pct': 77},
    'Jun-P2': {'chg': 78, 'sah': 79, 'chg_pct': 80},
    'Jul-P1': {'chg': 81, 'sah': 82, 'chg_pct': 83},
    'Jul-P2': {'chg': 84, 'sah': 85, 'chg_pct': 86},
    'Ago-P1': {'chg': 87, 'sah': 88, 'chg_pct': 89},
    'Ago-P2': {'chg': 90, 'sah': 91, 'chg_pct': 92},
    # Extensión Sep–Nov 2026 (FY27 calendar)
    'Sep-P1-26': {'chg': 93,  'sah': 94,  'chg_pct': 95},
    'Sep-P2-26': {'chg': 96,  'sah': 97,  'chg_pct': 98},
    'Oct-P1-26': {'chg': 99,  'sah': 100, 'chg_pct': 101},
    'Oct-P2-26': {'chg': 102, 'sah': 103, 'chg_pct': 104},
    'Nov-P1-26': {'chg': 105, 'sah': 106, 'chg_pct': 107},
}

# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def sf(val) -> float:
    """Conversión segura de celda a float. 0.0 si falla."""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace('%', '').strip())
    except (ValueError, TypeError):
        return 0.0


def parse_date_cell(val):
    """Maneja datetime / date / int / str provenientes de openpyxl."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Excel serial number (openpyxl a veces devuelve int cuando data_only=True
    # y la celda no tiene formato de fecha)
    if isinstance(val, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(val)
        except Exception:
            return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def classify_chg_cell(cell) -> str:
    """
    Examina el color de fondo de la celda y retorna:
      'hl'       → Hard Lock (sin color o blanco)
      'sl'       → Soft Lock (assumptions)
      'cascade'  → Cascadeo de horas
    """
    fill = cell.fill
    if fill is None or fill.fill_type in (None, 'none'):
        return 'hl'

    fg = fill.fgColor
    if fg is None:
        return 'hl'

    if fg.type == 'theme':
        # theme 4 = celeste "Assumptions R"
        return 'sl' if fg.theme == 4 else 'hl'

    if fg.type == 'rgb':
        rgb = fg.rgb
        if not rgb or rgb in ('00000000', 'FFFFFFFF'):
            return 'hl'
        if rgb in CASCADE_COLORS:
            return 'cascade'
        if rgb in SL_COLORS:
            return 'sl'
        # Color desconocido → conservador: HL, y se loggea
        log.warning(f'Color desconocido en {cell.coordinate}: {rgb!r} → clasificado como HL')
        return 'hl'

    return 'hl'


def extract_threaded_comments(excel_path: str) -> dict:
    """
    Lee los threaded comments (Office 365) directamente del ZIP del .xlsx.
    Retorna dict {cell_ref: text}, ej: {'U45': '+40hs PPA Ago P1 Client X'}.
    Si no hay threadedComments, retorna {}.
    """
    result = {}
    ns_re = re.compile(r'^\{.*?\}')

    try:
        with zipfile.ZipFile(excel_path) as zf:
            # Buscar todos los archivos threadedComments*.xml
            tc_files = [
                n for n in zf.namelist()
                if 'threadedComments' in n and n.endswith('.xml')
            ]
            for tc_file in tc_files:
                with zf.open(tc_file) as f:
                    tree = ET.parse(f)
                root = tree.getroot()
                ns = ns_re.match(root.tag)
                prefix = ns.group(0) if ns else ''
                # Cada <threadedComment ref="..."> puede tener múltiples <reply>
                for tc in root.iter(f'{prefix}threadedComment'):
                    ref = tc.get('ref', '')
                    texts = []
                    for r in tc.iter(f'{prefix}r'):
                        t_el = r.find(f'{prefix}t')
                        if t_el is not None and t_el.text:
                            texts.append(t_el.text)
                    if ref and texts:
                        result[ref] = ' '.join(texts)
    except Exception as e:
        log.warning(f'No se pudieron leer threaded comments: {e}')

    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def main(excel_path: str, dry_run: bool = False):
    try:
        import openpyxl
    except ImportError:
        log.error('openpyxl no instalado. Ejecutar: pip install "openpyxl>=3.1"')
        sys.exit(1)

    log.info(f'Leyendo Excel: {excel_path}')
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if SHEET_NAME:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active

    log.info(f'Hoja activa: {ws.title!r}')

    # Leer threaded comments
    comments = extract_threaded_comments(excel_path)
    log.info(f'Threaded comments encontrados: {len(comments)}')

    # Conectar a DB
    conn = await asyncpg.connect(
        host=settings.DB_HOST,
        port=settings.DB_PORT,
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        database=settings.DB_NAME,
        ssl='require',
    )

    # Períodos existentes en la DB
    db_periods = {
        r['period_name']
        for r in await conn.fetch('SELECT period_name FROM periods')
    }

    active_periods = {}
    skipped_periods = []
    for pname, cols in PERIOD_COLUMNS.items():
        if pname in db_periods:
            active_periods[pname] = cols
        else:
            skipped_periods.append(pname)

    if skipped_periods:
        log.warning(f'Períodos no encontrados en DB (se skipean): {skipped_periods}')
    log.info(f'Períodos activos: {list(active_periods.keys())}')

    # ---------------------------------------------------------------------------
    # Loop de filas
    # ---------------------------------------------------------------------------
    employees_data = []
    periods_data = []
    comment_log = []

    stats = {'hl': 0, 'sl': 0, 'cascade': 0, 'rows': 0, 'skipped_eid': 0}
    unknown_colors: dict = {}   # rgb → count

    for row_idx in range(DATA_ROW_START, DATA_ROW_END + 1):
        def cell(col): return ws.cell(row=row_idx, column=col)

        eid_val = cell(COL_EID).value
        if not eid_val or not str(eid_val).strip():
            stats['skipped_eid'] += 1
            continue

        eid = str(eid_val).strip().lower()
        stats['rows'] += 1

        # — Datos de empleado —
        location_raw = str(cell(COL_LOCATION).value or '').strip().upper()
        country, location = LOCATION_MAP.get(location_raw, ('AR', 'AR'))

        cl_val = cell(COL_CL).value
        cl = None
        if cl_val is not None:
            m = re.match(r'(\d+(?:\.\d+)?)', str(cl_val).strip())
            cl = float(m.group(1)) if m else None

        offering = str(cell(COL_OFFERING_LABEL).value or '').strip() or None
        career_track = str(cell(COL_OFFERING_REPORT).value or '').strip() or None
        client = str(cell(COL_CLIENT).value or '').strip() or None

        status_raw = str(cell(COL_STATUS).value or '').strip().lower()
        new_joiner = 'nj' in status_raw

        roll_on = parse_date_cell(cell(COL_ROLL_ON).value)
        roll_off = parse_date_cell(cell(COL_ROLL_OFF).value)
        first_avail = parse_date_cell(cell(COL_FIRST_AVAIL).value)
        next_client = str(cell(COL_NEXT_CLIENT).value or '').strip() or None

        employees_data.append({
            'eid': eid,
            'country': country,
            'location': location,
            'cl': cl,
            'offering': offering,
            'career_track': career_track,
            'new_joiner': new_joiner,
            'client': client,
            'roll_on': roll_on,
            'roll_off': roll_off,
            'first_available': first_avail,
            'next_client': next_client,
        })

        # — Datos de períodos —
        for pname, cols in active_periods.items():
            chg_cell = ws.cell(row=row_idx, column=cols['chg'])
            chg_val = sf(chg_cell.value)
            sah_val = sf(ws.cell(row=row_idx, column=cols['sah']).value)
            chg_pct_val = sf(ws.cell(row=row_idx, column=cols['chg_pct']).value)

            kind = classify_chg_cell(chg_cell)

            # Detectar colores desconocidos para estadísticas
            fill = chg_cell.fill
            if (fill and fill.fill_type not in (None, 'none') and
                    fill.fgColor and fill.fgColor.type == 'rgb'):
                rgb = fill.fgColor.rgb
                if rgb not in ('00000000', 'FFFFFFFF') and rgb not in SL_COLORS and rgb not in CASCADE_COLORS:
                    unknown_colors[rgb] = unknown_colors.get(rgb, 0) + 1

            chg_hl = chg_val if kind == 'hl' else 0.0
            chg_sl = chg_val if kind == 'sl' else 0.0
            chg_cascade = chg_val if kind == 'cascade' else 0.0

            stats[kind] += 1

            chg_pct_hl = (chg_hl / sah_val * 100) if sah_val > 0 else 0.0
            chg_pct_sl = (chg_sl / sah_val * 100) if sah_val > 0 else 0.0

            # Buscar comment en esta celda
            cell_ref = chg_cell.coordinate
            comment_text = comments.get(cell_ref)
            if comment_text:
                comment_log.append({
                    'eid': eid,
                    'period': pname,
                    'cell_ref': cell_ref,
                    'text': comment_text,
                })

            periods_data.append({
                'eid': eid,
                'period_name': pname,
                'chg': chg_val,
                'sah': sah_val,
                'chg_pct': chg_pct_val,
                'chg_hl': chg_hl,
                'chg_sl': chg_sl,
                'chg_cascadeadas': chg_cascade,
                'chg_pct_hl': chg_pct_hl,
                'chg_pct_sl': chg_pct_sl,
            })

    # ---------------------------------------------------------------------------
    # Diagnóstico
    # ---------------------------------------------------------------------------
    log.info('--- DIAGNÓSTICO ---')
    log.info(f'Empleados procesados: {stats["rows"]} (filas vacías skipeadas: {stats["skipped_eid"]})')
    log.info(f'Filas de período generadas: {len(periods_data)}')
    log.info(f'Celdas CHG clasificadas → HL: {stats["hl"]}, SL: {stats["sl"]}, Cascade: {stats["cascade"]}')
    if unknown_colors:
        log.warning(f'Colores desconocidos encontrados: {unknown_colors}')
    log.info(f'Comments encontrados en celdas CHG: {len(comment_log)}')

    # Guardar log de comments
    log_path = Path('excel_comments.log')
    with open(log_path, 'w', encoding='utf-8') as lf:
        lf.write('eid\tperiod\tcell_ref\ttext\n')
        for entry in comment_log:
            lf.write(f'{entry["eid"]}\t{entry["period"]}\t{entry["cell_ref"]}\t{entry["text"]}\n')
    log.info(f'Log de comments guardado en: {log_path.resolve()}')

    if dry_run:
        log.info('--- DRY RUN: no se escribe en la DB ---')
        await conn.close()
        return

    # ---------------------------------------------------------------------------
    # Transacción única → upsert employees + forecast_update + forecast_periods
    # ---------------------------------------------------------------------------
    log.info('Escribiendo en la DB...')
    emp_inserted = emp_updated = fu_upserted = fp_upserted = 0

    async with conn.transaction():
        for emp in employees_data:
            eid = emp['eid']
            exists = await conn.fetchval(
                'SELECT eid FROM employees WHERE eid=$1', eid
            )
            if exists:
                await conn.execute("""
                    UPDATE employees SET
                        country  = COALESCE($2, country),
                        location = COALESCE($3, location),
                        cl       = COALESCE($4, cl),
                        offering = COALESCE($5, offering),
                        career_track = COALESCE($6, career_track),
                        new_joiner   = $7,
                        charge       = TRUE
                    WHERE eid = $1
                """, eid,
                    emp['country'], emp['location'],
                    emp['cl'], emp['offering'], emp['career_track'],
                    emp['new_joiner'],
                )
                emp_updated += 1
            else:
                await conn.execute("""
                    INSERT INTO employees
                        (eid, name, country, location, cl, offering, career_track,
                         new_joiner, active, charge)
                    VALUES ($1, $1, $2, $3, $4, $5, $6, $7, TRUE, TRUE)
                """, eid,
                    emp['country'], emp['location'],
                    emp['cl'], emp['offering'], emp['career_track'],
                    emp['new_joiner'],
                )
                emp_inserted += 1

            # forecast_update
            await conn.execute("""
                INSERT INTO forecast_update
                    (eid, client, roll_on, roll_off, first_available, next_client, updated_at)
                VALUES ($1, $2, $3, $4, $5, $6, NOW())
                ON CONFLICT (eid) DO UPDATE SET
                    client          = COALESCE($2, forecast_update.client),
                    roll_on         = COALESCE($3, forecast_update.roll_on),
                    roll_off        = COALESCE($4, forecast_update.roll_off),
                    first_available = COALESCE($5, forecast_update.first_available),
                    next_client     = COALESCE($6, forecast_update.next_client),
                    updated_at      = NOW()
                -- chargeability_pct no se toca: lo gestiona la app
            """, eid,
                emp['client'], emp['roll_on'], emp['roll_off'],
                emp['first_available'], emp['next_client'],
            )
            fu_upserted += 1

        # forecast_periods — el fix clave: incluir chg_cascadeadas en ON CONFLICT
        for fp in periods_data:
            await conn.execute("""
                INSERT INTO forecast_periods
                    (eid, period_name, chg, sah, chg_pct,
                     chg_hl, chg_sl, chg_cascadeadas,
                     absence_hours, chg_pct_sl, chg_pct_hl)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, 0, $9, $10)
                ON CONFLICT (eid, period_name) DO UPDATE SET
                    chg             = $3,
                    sah             = $4,
                    chg_pct         = $5,
                    chg_hl          = $6,
                    chg_sl          = $7,
                    chg_cascadeadas = $8,
                    chg_pct_sl      = $9,
                    chg_pct_hl      = $10
                    -- absence_hours NO se toca: lo maneja el pipeline de ausencias
            """,
                fp['eid'], fp['period_name'],
                fp['chg'], fp['sah'], fp['chg_pct'],
                fp['chg_hl'], fp['chg_sl'], fp['chg_cascadeadas'],
                fp['chg_pct_sl'], fp['chg_pct_hl'],
            )
            fp_upserted += 1

    log.info('--- RESULTADO ---')
    log.info(f'Employees insertados:  {emp_inserted}')
    log.info(f'Employees actualizados: {emp_updated}')
    log.info(f'Forecast update upserts: {fu_upserted}')
    log.info(f'Forecast periods upserts: {fp_upserted}')

    await conn.close()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='Carga el Excel de forecast a la DB preservando colores de celda.'
    )
    parser.add_argument(
        'excel_path',
        nargs='?',
        help='Path al archivo .xlsx. Si se omite, busca el primero en el directorio actual.',
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Lee el Excel y muestra diagnóstico sin escribir en la DB.',
    )
    args = parser.parse_args()

    if args.excel_path:
        path = args.excel_path
    else:
        xlsx_files = list(Path('.').glob('*.xlsx'))
        if not xlsx_files:
            log.error('No se encontró ningún archivo .xlsx en el directorio actual.')
            sys.exit(1)
        path = str(xlsx_files[0])
        log.info(f'Excel encontrado automáticamente: {path}')

    asyncio.run(main(path, dry_run=args.dry_run))
