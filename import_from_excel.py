"""
import_from_excel.py
--------------------
Lee el Excel de Forecast S&P y carga chargeability_blocks + ppa_log en Azure.

Distingue HL / SL / PPA por color de celda en las columnas CHG:

    sin relleno / 00000000  -> HL   (effective)
    FFFFC7CE, FFFFCCCC      -> SL   (assumption)  [rojo]
    FF44B3E1, FFC0E6F5      -> SL   (assumption)  [celeste]
    FFCC66FF                -> PPA  (cascadeo)    [lila] - horas del comentario
    FFFFEB9C                -> sin clasificar (se reporta, no se carga)

Uso:
    python import_from_excel.py "<excel>" --dry-run     # solo reporta
    python import_from_excel.py "<excel>"               # escribe en la DB
"""

import argparse
import asyncio
import os
import re
import logging
from collections import Counter

import asyncpg
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S')
log = logging.getLogger(__name__)

AZURE = dict(
    host=os.getenv('DB_HOST'),
    port=int(os.getenv('DB_PORT', 5432)),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    database=os.getenv('DB_NAME'),
    ssl='require',
)

PERIODOS_A_LIMPIAR = []

SHEET = 'Forecast Update'
# La fila 2 repite los periodos: cols 21-42 son el bloque VIEJO (Sep-Dic 2025)
# y cols 46-105 el bloque VIGENTE (Ene 2026 - Nov 2026). Solo leemos el vigente,
# si no cargamos datos de 2025 en los periodos de 2026.
MIN_PERIOD_COL = 46
HEADER_ROW = 2
SUBHEADER_ROW = 10   # fila con 'CHG' / 'SAH' / 'CHG (%)'
FIRST_DATA_ROW = 11
EID_COL = 2

# Leyenda real del Excel -> hoja 'Forecast Update', celdas B93:B96
#   B93 FFFFEB9C  Assumptions ISG Assessment
#   B94 FFFFCCCC  Assumptions No R
#   B95 THEME4    Assumptions R
#   B96 FFCC66FF  Cascadeo de hs
COLOR_KIND = {
    'NOFILL':        'HL',
    '00000000':      'HL',
    'THEME0+0.0':    'HL',     # blanco de tema = sin color
    'THEME0+-0.15':  'HL',

    'FFFFEB9C':      'SL',     # Assumptions ISG Assessment
    'FFFFCCCC':      'SL',     # Assumptions No R
    'FFFFC7CE':      'SL',     # rosa variante de No R
    'THEME4+0.8':    'SL',     # Assumptions R

    'FFCC66FF':      'PPA',    # Cascadeo de hs

    'FF44B3E1':      'REVISAR',  # celeste - probable LOA (sick days)
    'FFC0E6F5':      'REVISAR',
    'FF00B0F0':      'REVISAR',
}

# Subtipo de assumption -> se guarda en chargeability_blocks.assumption_kind
# para que el front pueda pintar la celda como el Excel.
COLOR_SUBKIND = {
    'FFFFEB9C': 'isg_assessment',   # Assumptions ISG Assessment (90%)
    'FFFFCCCC': 'no_r',             # Assumptions No R (2p 0% luego 50%)
    'FFFFC7CE': 'no_r',             # variante rosa de No R
    'THEME4+0.8': 'r',              # Assumptions R (1p 0%, 2p 75%, luego 100%)
}

PERIOD_LABEL_MAP = {
    'AGO P1': 'Ago-P1', 'AGO P2': 'Ago-P2',
    'SEP P1': 'Sep-P1', 'SEP P2': 'Sep-P2',
    'OCT P1': 'Oct-P1', 'OCT P2': 'Oct-P2',
    'NOV P1': 'Nov-P1', 'NOV P2': 'Nov-P2',
    'DIC P1': 'Dic-P1', 'DIC P2': 'Dic-P2',
    'ENE P1': 'Ene-P1', 'ENE P2': 'Ene-P2',
    'FEB P1': 'Feb-P1', 'FEB P2': 'Feb-P2',
    'MAR P1': 'Mar-P1', 'MAR P2': 'Mar-P2',
    'ABR P1': 'Abr-P1', 'ABR P2': 'Abr-P2',
    'MAY P1': 'May-P1', 'MAY P2': 'May-P2',
    'JUN P1': 'Jun-P1', 'JUN P2': 'Jun-P2',
    'JUL P1': 'Jul-P1', 'JUL P2': 'Jul-P2',
}

MONTH_ALIASES = {
    'ENE': 'Ene', 'ENERO': 'Ene',
    'FEB': 'Feb', 'FEBRERO': 'Feb',
    'MAR': 'Mar', 'MARZO': 'Mar',
    'ABR': 'Abr', 'ABRIL': 'Abr',
    'MAY': 'May', 'MAYO': 'May',
    'JUN': 'Jun', 'JUNIO': 'Jun',
    'JUL': 'Jul', 'JULIO': 'Jul',
    'AGO': 'Ago', 'AGOSTO': 'Ago',
    'SEP': 'Sep', 'SEPT': 'Sep', 'SEPTIEMBRE': 'Sep',
    'OCT': 'Oct', 'OCTUBRE': 'Oct',
    'NOV': 'Nov', 'NOVIEMBRE': 'Nov',
    'DIC': 'Dic', 'DICIEMBRE': 'Dic',
}

# El comentario tiene que hablar de cascadeo para contarlo como PPA
RE_ES_PPA = re.compile(r'\b(ppa|cascade|cascade[oa]|carga|cedid)', re.IGNORECASE)
# "7*8" / "3 * 8"  -> 56 / 24
RE_MULT = re.compile(r'(\d+)\s*\*\s*(\d+)')
# "+40hs" / "64 horas"
RE_HOURS = re.compile(r'([+-]?\s*\d+(?:[.,]\d+)?)\s*(?:hs|h|horas)\b', re.IGNORECASE)
# "PPA 72 Abr P1" -> numero suelto pegado a la palabra PPA
RE_BARE = re.compile(r'\bPPA\s*(?:de\s*)?(\d+(?:[.,]\d+)?)\b', re.IGNORECASE)
RE_PERIOD = re.compile(
    r'\b(' + '|'.join(sorted(MONTH_ALIASES, key=len, reverse=True)) + r')\.?\s*(?:P\s*)?([12])\b',
    re.IGNORECASE,
)


def cell_color(cell) -> str:
    """Color de relleno. Soporta rgb, theme e indexed (theme era el bug: los
    'Assumptions R' usan theme4+0.8 y se leian como sin relleno)."""
    f = cell.fill
    if not f or not f.patternType:
        return 'NOFILL'
    fg = f.fgColor
    if not fg:
        return 'NOFILL'
    try:
        if fg.type == 'rgb' and fg.rgb and str(fg.rgb).upper() != '00000000':
            return str(fg.rgb).upper()
        if fg.type == 'theme':
            return f'THEME{fg.theme}+{round(float(fg.tint), 3)}'
        if fg.type == 'indexed':
            return f'INDEXED{fg.indexed}'
    except Exception:
        pass
    return 'NOFILL'


def clean_comment(cell):
    if not cell.comment or not cell.comment.text:
        return None
    txt = cell.comment.text
    if 'Comment:' in txt:
        txt = txt.split('Comment:', 1)[1]
    txt = re.sub(r'\[Threaded comment\].*?Learn more:\s*\S+', '', txt, flags=re.DOTALL)
    return ' '.join(txt.split()).strip() or None


def parse_ppa(comment):
    """
    '+40hs PPA Dic P1'   -> (40.0, 'Dic-P1')
    'PPA AGO P1 7*8'     -> (56.0, 'Ago-P1')
    'PPA 72 Abr P1'      -> (72.0, 'Abr-P1')
    '252 hs total Ago'   -> (None, None)   # no menciona PPA, no es cascadeo
    """
    if not comment:
        return None, None

    # 1) multiplicacion primero: "3*4 hs" son 12, no 4
    hours = None
    m = RE_MULT.search(comment)
    if m:
        hours = float(m.group(1)) * float(m.group(2))
    if hours is None:
        m = RE_HOURS.search(comment)
        if m:
            hours = float(m.group(1).replace(' ', '').replace(',', '.'))
    if hours is None:
        m = RE_BARE.search(comment)
        if m:
            hours = float(m.group(1).replace(',', '.'))
    if hours is None:
        return None, None

    # 2) el periodo origen tiene que venir con P1/P2. Sin eso no es una
    #    transferencia ("252 hs total Ago" es una nota, no un cascadeo).
    p = RE_PERIOD.search(comment)
    if not p:
        return None, None
    period = f"{MONTH_ALIASES[p.group(1).upper()]}-P{p.group(2)}"
    return abs(hours), period


def num(cell):
    v = cell.value
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def build_triplets(ws):
    """
    Arma los tripletes (chg, sah, pct) leyendo la fila de subcabeceras.
    No se puede asumir que la columna del label sea la de CHG: en ENE P1 el
    label esta corrido y cae sobre SAH, lo que daba porcentajes de 8000%.
    """
    tipo = {}
    for cell in ws[SUBHEADER_ROW]:
        if not isinstance(cell.value, str):
            continue
        h = cell.value.strip().upper()
        # el '%' puede venir en cualquier posicion: 'CHG (%)', 'CHG Jun P2 (%)'
        if '%' in h:
            tipo[cell.column] = 'PCT'
        elif h.startswith('CHG'):
            tipo[cell.column] = 'CHG'
        elif h.startswith('SAH'):
            tipo[cell.column] = 'SAH'

    triplets = []
    for col in sorted(tipo):
        if tipo[col] != 'CHG':
            continue
        sah = next((c for c in sorted(tipo) if c > col and tipo[c] == 'SAH'), None)
        pct = next((c for c in sorted(tipo) if c > col and tipo[c] == 'PCT'), None)
        if sah and pct and sah < pct:
            triplets.append((col, sah, pct))
    return triplets


def triplet_for(triplets, label_col):
    """Devuelve el triplete al que pertenece la columna del label del periodo."""
    for chg, sah, pct in triplets:
        if chg <= label_col <= pct:
            return chg, sah, pct
    # fallback: el triplete cuyo CHG este mas cerca
    if not triplets:
        return None
    return min(triplets, key=lambda t: abs(t[0] - label_col))


def parse_excel(path):
    log.info(f'Leyendo {path}')
    wb = load_workbook(path, read_only=False, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"No existe la hoja '{SHEET}'. Hojas: {wb.sheetnames}")
    ws = wb[SHEET]

    period_cols, descartados = {}, {}
    for cell in ws[HEADER_ROW]:
        if isinstance(cell.value, str):
            key = cell.value.strip().upper()
            if key not in PERIOD_LABEL_MAP:
                continue
            period = PERIOD_LABEL_MAP[key]
            if cell.column < MIN_PERIOD_COL:
                descartados.setdefault(period, []).append(cell.column)
                continue
            period_cols[period] = cell.column   # ultima aparicion gana

    solo_viejas = {p: c for p, c in descartados.items() if p not in period_cols}
    if solo_viejas:
        log.warning(f'Sin datos vigentes en el Excel (solo bloque 2025): '
                    f'{sorted(solo_viejas)} -> se limpian de la DB')

    if not period_cols:
        raise SystemExit('No se detectaron columnas de periodo en la fila 2.')
    log.info(f'Periodos vigentes detectados: {len(period_cols)} -> {sorted(period_cols)}')
    PERIODOS_A_LIMPIAR.clear()
    PERIODOS_A_LIMPIAR.extend(sorted(solo_viejas))

    triplets = build_triplets(ws)
    log.info(f'Tripletes CHG/SAH/CHG% detectados: {len(triplets)}')

    # mapea periodo -> (col_chg, col_sah, col_pct) usando las subcabeceras
    period_tri = {}
    for period, lcol in period_cols.items():
        t = triplet_for(triplets, lcol)
        if t is None:
            log.warning(f'{period}: no encontre triplete, se omite')
            continue
        period_tri[period] = t
        if t[0] != lcol:
            log.warning(f'{period}: el label esta en col {lcol} pero CHG es col {t[0]} (corregido)')

    blocks, ppas, unknown, ppa_sin_dato = [], [], [], []
    sah_map = {}
    fuera_de_rango = []
    kinds = Counter()

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        eid = ws.cell(row=r, column=EID_COL).value
        if not eid or not isinstance(eid, str) or eid.startswith('='):
            continue
        eid = eid.strip()

        for period, (ccol, scol, pcol) in period_tri.items():
            chg_cell = ws.cell(row=r, column=ccol)
            sah_cell = ws.cell(row=r, column=scol)
            pct_cell = ws.cell(row=r, column=pcol)

            color = cell_color(chg_cell)
            kind = COLOR_KIND.get(color, 'REVISAR')
            kinds[kind] += 1

            chg = num(chg_cell)
            sah = num(sah_cell)
            comment = clean_comment(chg_cell)

            # El SAH se guarda aparte de los bloques porque aplica igual cuando
            # el CHG es 0 o el color no se puede clasificar, casos en los que el
            # bloque se descarta pero las horas disponibles siguen siendo validas.
            if sah is not None:
                sah_map[(eid, period)] = sah

            if kind == 'PPA':
                hours, from_period = parse_ppa(comment)
                if hours is None:
                    ppa_sin_dato.append((eid, period, chg_cell.coordinate, comment))
                else:
                    ppas.append({
                        'eid': eid, 'to_period': period,
                        'from_period': from_period, 'hours': hours,
                        'reason': (comment or 'PPA importado del Excel')[:200],
                    })
                kind = 'HL'   # el CHG de una celda con PPA sigue siendo HL

            if kind == 'REVISAR':
                unknown.append((eid, period, chg_cell.coordinate, cell_color(chg_cell)))
                continue

            if chg is None or sah is None:
                continue

            # El Excel ya trae el CHG% calculado (como fraccion: 1 = 100%).
            # Usarlo es mas fiel que recalcular, y evita divisiones raras.
            pct_raw = num(pct_cell)
            if pct_raw is not None:
                pct = round(pct_raw * 100, 2)
            elif sah > 0:
                pct = round(chg / sah * 100, 2)
            else:
                pct = 100.0 if chg > 0 else 0.0

            if pct > 150:
                fuera_de_rango.append((eid, period, chg_cell.coordinate, chg, sah, pct))
            es_sl = (kind == 'SL')
            blocks.append({
                'eid': eid, 'period_name': period,
                'chargeability_pct': pct,
                'scenario_type': 'assumption' if es_sl else 'effective',
                'assumption_kind': COLOR_SUBKIND.get(color) if es_sl else None,
            })

    log.info(f'Celdas por tipo: {dict(kinds)}')
    log.info(f'Bloques: {len(blocks)} | PPA con horas: {len(ppas)} | '
             f'PPA sin dato: {len(ppa_sin_dato)} | Para revisar: {len(unknown)}')
    if fuera_de_rango:
        log.warning(f'{len(fuera_de_rango)} celdas con CHG% > 150% (revisar en el Excel)')
        for x in fuera_de_rango[:12]:
            log.warning(f'   {x[0]} {x[1]} {x[2]}  chg={x[3]} sah={x[4]} -> {x[5]}%')

    log.info(f'SAH leidos del Excel: {len(sah_map)} pares (eid, periodo)')

    return blocks, ppas, unknown, ppa_sin_dato, sah_map


async def write_db(blocks, ppas, sah_map):
    conn = await asyncpg.connect(**AZURE)
    try:
        await conn.execute(
            'ALTER TABLE chargeability_blocks '
            'ADD COLUMN IF NOT EXISTS assumption_kind VARCHAR(30)')

        periods = {r['period_name']: (r['start_date'], r['end_date'])
                   for r in await conn.fetch('SELECT period_name, start_date, end_date FROM periods')}
        eids = {r['eid'] for r in await conn.fetch('SELECT eid FROM employees WHERE active = TRUE')}
        log.info(f'DB: {len(periods)} periodos, {len(eids)} empleados activos')

        if PERIODOS_A_LIMPIAR:
            borrados = 0
            for pn in PERIODOS_A_LIMPIAR:
                r = await conn.execute(
                    'DELETE FROM chargeability_blocks WHERE period_name = $1', pn)
                borrados += int(r.split()[-1]) if r.split()[-1].isdigit() else 0
            log.info(f'Limpieza de periodos sin dato vigente {PERIODOS_A_LIMPIAR} '
                     f'-> {borrados} bloques borrados')

        ok = skip_eid = skip_period = 0
        validos = []
        for b in blocks:
            if b['eid'] not in eids:
                skip_eid += 1
                continue
            if b['period_name'] not in periods:
                skip_period += 1
                continue
            if b['chargeability_pct'] <= 0:
                continue
            start, end = periods[b['period_name']]
            validos.append((b['eid'], b['period_name'], float(b['chargeability_pct']),
                            b['scenario_type'], b.get('assumption_kind'), start, end))
            ok += 1

        # Escritura en lote. Antes era un DELETE+INSERT por fila (~2500
        # round-trips) y Azure cortaba la conexion a mitad de camino.
        # El DELETE va por (eid, periodo) sin filtrar scenario_type: si una
        # celda cambia de HL a SL entre corridas, filtrar por tipo dejaba viva
        # la fila anterior y quedaban las dos (el CHG Neto se duplicaba).
        claves = sorted({(v[0], v[1]) for v in validos})
        log.info(f'Limpiando {len(claves)} combinaciones (eid, periodo)...')
        async with conn.transaction():
            await conn.executemany(
                'DELETE FROM chargeability_blocks WHERE eid = $1 AND period_name = $2',
                claves)
            log.info(f'Insertando {len(validos)} bloques...')
            await conn.executemany(
                'INSERT INTO chargeability_blocks '
                '(eid, period_name, chargeability_pct, scenario_type, '
                ' assumption_kind, start_date, end_date, created_at) '
                'VALUES ($1, $2, $3, $4, $5, $6, $7, NOW())',
                validos)

        log.info(f'chargeability_blocks -> {ok} insertados | '
                 f'{skip_eid} EID desconocido | {skip_period} periodo desconocido')

        # El SAH del Excel es la fuente de verdad. Antes forecast_periods.sah
        # conservaba lo que habia cargado load_periods.py una sola vez, y como
        # el CHG se deriva de ese SAH, ninguno de los dos coincidia con el Excel.
        sah_rows = [(e, pn, float(v)) for (e, pn), v in sorted(sah_map.items())
                    if e in eids and pn in periods]
        excel_periods = sorted({pn for _e, pn, _v in sah_rows})
        if sah_rows:
            log.info(f'Actualizando SAH de {len(sah_rows)} filas en '
                     f'{len(excel_periods)} periodos...')
            async with conn.transaction():
                await conn.executemany(
                    'INSERT INTO forecast_periods (eid, period_name, sah) '
                    'VALUES ($1, $2, $3) '
                    'ON CONFLICT (eid, period_name) DO UPDATE SET sah = EXCLUDED.sah',
                    sah_rows)

                # Rederiva el CHG con el SAH nuevo. Misma formula que
                # recalculate_service.recalculate_employee, en un solo statement.
                await conn.execute(
                    """
                    WITH totals AS (
                        SELECT eid, period_name,
                               COALESCE(SUM(chargeability_pct)
                                   FILTER (WHERE scenario_type = 'effective'),  0) AS hl_pct,
                               COALESCE(SUM(chargeability_pct)
                                   FILTER (WHERE scenario_type = 'assumption'), 0) AS sl_pct
                        FROM chargeability_blocks
                        WHERE period_name = ANY($1)
                        GROUP BY eid, period_name
                    )
                    UPDATE forecast_periods fp
                    SET chg_pct_hl = t.hl_pct,
                        chg_pct_sl = t.sl_pct,
                        chg_hl     = ROUND(fp.sah * t.hl_pct / 100.0),
                        chg_sl     = ROUND(fp.sah * t.sl_pct / 100.0),
                        chg        = ROUND(fp.sah * (t.hl_pct + t.sl_pct) / 100.0)
                    FROM totals t
                    WHERE fp.eid = t.eid AND fp.period_name = t.period_name
                    """,
                    excel_periods)

                # Sin bloques el empleado no tiene horas cargables en ese
                # periodo. recalculate_service deja 0; si no se hace explicito,
                # queda el CHG viejo de la corrida anterior.
                await conn.execute(
                    """
                    UPDATE forecast_periods fp
                    SET chg_pct_hl = 0, chg_pct_sl = 0,
                        chg_hl = 0, chg_sl = 0, chg = 0
                    WHERE fp.period_name = ANY($1)
                      AND NOT EXISTS (
                          SELECT 1 FROM chargeability_blocks cb
                          WHERE cb.eid = fp.eid AND cb.period_name = fp.period_name
                      )
                    """,
                    excel_periods)
            log.info('forecast_periods.sah <- Excel, CHG rederivado')

        exists = await conn.fetchval("SELECT to_regclass('public.ppa_log')")
        if not exists:
            log.warning('No existe ppa_log, se omite el cascadeo.')
            return

        existentes = {(r['eid'], r['from_period'], r['to_period'], float(r['hours']))
                      for r in await conn.fetch(
                          'SELECT eid, from_period, to_period, hours FROM ppa_log')}
        nuevos = []
        for p in ppas:
            if p['eid'] not in eids or not p['from_period']:
                continue
            k = (p['eid'], p['from_period'], p['to_period'], float(p['hours']))
            if k in existentes:
                continue
            existentes.add(k)
            nuevos.append((p['eid'], p['from_period'], p['to_period'],
                           float(p['hours']), p['reason']))
        if nuevos:
            async with conn.transaction():
                await conn.executemany(
                    'INSERT INTO ppa_log (eid, from_period, to_period, hours, reason, created_at) '
                    'VALUES ($1, $2, $3, $4, $5, NOW())', nuevos)
        n = len(nuevos)
        log.info(f'ppa_log -> {n} filas nuevas')
    finally:
        await conn.close()


def main():
    ap = argparse.ArgumentParser()
    global MIN_PERIOD_COL

    ap.add_argument('excel')
    ap.add_argument('--dry-run', action='store_true', help='no escribe en la DB')
    # El bloque vigente se corre a la derecha cada mes. Sin esta opcion habia
    # que editar el codigo, y si quedaba viejo el importer descartaba periodos
    # buenos en silencio.
    ap.add_argument('--min-period-col', type=int, default=MIN_PERIOD_COL,
                    help=f'primera columna del bloque vigente (default {MIN_PERIOD_COL}). '
                         'Revisa el log "Periodos vigentes detectados" para confirmar.')
    args = ap.parse_args()

    MIN_PERIOD_COL = args.min_period_col

    if not os.path.exists(args.excel):
        raise SystemExit(f'No se encontro el archivo: {args.excel}')

    blocks, ppas, unknown, ppa_sin_dato, sah_map = parse_excel(args.excel)

    hl = sum(1 for b in blocks if b['scenario_type'] == 'effective')
    sl = sum(1 for b in blocks if b['scenario_type'] == 'assumption')
    print(f'\n  HL (effective) : {hl}')
    print(f'  SL (assumption): {sl}')
    print(f'  PPA / cascadeo : {len(ppas)}')
    sub = Counter(b['assumption_kind'] for b in blocks if b['scenario_type'] == 'assumption')
    for k, n in sorted(sub.items(), key=lambda x: -x[1]):
        print(f'      SL {k or "sin subtipo"}: {n}')

    if ppas:
        print('\n  --- PPA detectados (primeros 15) ---')
        for p in ppas[:15]:
            print(f'    {p["eid"]:28} {str(p["from_period"] or "?"):8} -> {p["to_period"]:8} {p["hours"]:>6.0f} hs')

    if ppa_sin_dato:
        print(f'\n  --- Celdas lila SIN horas parseables ({len(ppa_sin_dato)}) - revisar a mano ---')
        for eid, period, coord, com in ppa_sin_dato[:15]:
            print(f'    {eid:28} {period:8} {coord:6} {com or "(sin comentario)"}')

    if unknown:
        print(f'\n  --- Color sin clasificar / a revisar ({len(unknown)}) - NO se cargaron ---')
        for eid, period, coord, color in unknown[:15]:
            print(f'    {eid:28} {period:8} {coord:6} {color}')

    if args.dry_run:
        print('\n  DRY RUN: no se escribio nada en la DB.\n')
        return

    if not all([AZURE['host'], AZURE['user'], AZURE['password'], AZURE['database']]):
        raise SystemExit('Faltan variables de DB en .env')

    asyncio.run(write_db(blocks, ppas, sah_map))
    print('\n  Listo. Ahora corre: python create_daily_hours.py\n')


if __name__ == '__main__':
    main()